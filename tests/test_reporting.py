import unittest
from datetime import datetime
from decimal import Decimal

from extraction import parse_data
from persisting import persist_results
from reporting import create_extract
from domain import YearMonth, TradeType, TradePartsWithinMonth, MonthPartitionedTrades, TradeCyclePerCompany, \
    CapitalGainLinesPerCompany, QuantitatedTradeAction, QuantitatedTradeActions, get_year_month
from tests.data import sell_action1
from transformation import split_by_months

test_dict1 = {("2022", "01"), ("2021", "12"), ("2021", "02")}
test_dict2 = ["202201", "202112", "202102"]
date_time1 = datetime.strptime("2021-05-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
date_time2 = datetime.strptime("2022-05-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
date_time3 = datetime.strptime("2021-01-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
date_time4 = datetime.strptime("2021-12-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
test_dict4 = [YearMonth(date_time1.year, date_time1.month), YearMonth(date_time2.year, date_time2.month), YearMonth(date_time3.year, date_time3.month), YearMonth(date_time4.year, date_time4.month)]


class MyTestCase(unittest.TestCase):

    def test_sorting(self):
        print(str(test_dict1))
        print(str(sorted(test_dict1)))
        print(str(test_dict2))
        print(str(sorted(test_dict2)))
        print(str(test_dict4))
        print(str(sorted(test_dict4)))
        self.assertEqual(1, 1)

    def test_partitioning(self):
        trades_within_month1 = TradePartsWithinMonth()
        trades_within_month1.push_trade_part(Decimal(1), sell_action1)
        month_partitioned_trades1: MonthPartitionedTrades = {get_year_month(sell_action1.date_time): trades_within_month1}

        actions: QuantitatedTradeActions = [QuantitatedTradeAction(quantity=Decimal(1.0), action=sell_action1)]
        actual: MonthPartitionedTrades = split_by_months(actions, TradeType.SELL)
        print(actual)
        self.assertEqual(actual, month_partitioned_trades1)

    # https://stackoverflow.com/questions/52089716/comparing-two-excel-files-using-openpyxl
    def test_reporting(self):
        from pathlib import Path

        expected = Path('resources', 'extract.xlsx')
        source = Path('resources', 'shares.csv')
        destination = Path('resources', 'tmp.xlsx')
        leftover = Path('resources', 'tmp_leftover.xlsx')

        trade_actions_per_company: TradeCyclePerCompany = parse_data(source)
        capital_gain_lines_per_company: CapitalGainLinesPerCompany = create_extract(trade_actions_per_company)
        persist_results(destination, leftover, capital_gain_lines_per_company)

        from openpyxl import load_workbook
        workbook1 = load_workbook(expected, data_only=False)
        workbook2 = load_workbook(destination, data_only=False)

        sheet1 = workbook1.active
        sheet2 = workbook2.active
        self.assertEqual(sheet1.title, sheet2.title)

        min_row1: int = sheet1.min_row
        min_row2: int = sheet2.min_row
        self.assertEqual(min_row1, min_row2)

        max_row1: int = sheet1.max_row
        max_row2: int = sheet2.max_row
        self.assertEqual(max_row1, max_row2)

        min_column1: int = sheet1.min_column
        min_column2: int = sheet2.min_column
        self.assertEqual(min_column1, min_column2)

        max_column1: int = sheet1.max_column
        max_column2: int = sheet2.max_column
        self.assertEqual(max_column1, max_column2)

        for i in range(min_row1, max_row1):
            for j in range(min_column1, max_column1):
                expected = sheet1.cell(i, j).value
                actual = sheet2.cell(i, j).value
                self.assertEqual(expected, actual,
                                 "Values at row " + str(i) + " and column " + str(j) + " differs:\n" +
                                 str(expected) + " expected and\n" + str(actual) + " actual\n")


if __name__ == '__main__':
    unittest.main()
