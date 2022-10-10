import os
import openpyxl
from typing import Union, List, Dict

from openpyxl.worksheet.worksheet import Worksheet

from config import Config, read_config, ConversionRate
from domain import CapitalGainLinesPerCompany, InciRecord

first_header = ["Beneficiary", "Country of Source", "SALE", "", "", "PURCHASE", "", "",
                "WITHOLDING TAX", "", "Expenses incurred with obtaining the capital gains", "",
                "Symbol", "Currency", "Sale amount", "Buy amount", "Expenses amount"]
second_header = ["", "", "Month ", "Year", "Amount", "Month ", "Year", "Amount", "Country", "Amount",
                 "", "", "", "", "", "", ""]
currency_header = ["Base/target", "Rate"]

last_column: int = max(len(first_header), len(second_header))


def safe_remove_file(path: Union[str, os.PathLike[str]]):
    if os.path.exists(path):
        os.remove(path)


def persist_as_query(path: Union[str, os.PathLike[str]], incis:  List[InciRecord]):
    number_format = '0.000000'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporting"

    config: Config = read_config()
    exchange_rates: Dict[str, str] = create_currency_table(worksheet, last_column + 2, 1, config)

    for i in range(len(first_header)):
        worksheet.cell(1, i + 1, first_header[i])
        worksheet.cell(2, i + 1, second_header[i])

    start_column = 3
    line_number = 3
    for currency_company, capital_gain_lines in incis.items():
        currency = currency_company[0]
        symbol = currency_company[1]
        for line in capital_gain_lines:
            assert currency == line.get_currency()
            idx = start_column
            c = worksheet.cell(line_number, start_column, line.get_sell_date().get_month_name())
            idx += 1
            c = worksheet.cell(line_number, idx, line.get_sell_date().year)
            idx += 1
            c = worksheet.cell(line_number, idx, "=" + exchange_rates[currency] + "*(" + line.get_sell_amount() + ")")
            idx += 1
            c = worksheet.cell(line_number, idx, line.get_buy_date().get_month_name())
            idx += 1
            c = worksheet.cell(line_number, idx, line.get_buy_date().year)
            idx += 1
            c = worksheet.cell(line_number, idx, "=" + exchange_rates[currency] + "*(" + line.get_buy_amount() + ")")
            idx += 3
            c = worksheet.cell(line_number, idx, "=" + exchange_rates[currency] + "*(" + line.get_expense_amount() +
                               ")")
            idx += 2
            c = worksheet.cell(line_number, idx, symbol)
            idx += 1
            c = worksheet.cell(line_number, idx, currency)
            idx += 1
            c = worksheet.cell(line_number, idx, line.get_sell_amount())
            c.number_format = number_format
            idx += 1
            c = worksheet.cell(line_number, idx, line.get_buy_amount())
            c.number_format = number_format
            idx += 1
            c = worksheet.cell(line_number, idx, line.get_expense_amount())
            c.number_format = number_format
            line_number += 1

    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2


    safe_remove_file(path)
    workbook.save(path)
    workbook.close()
