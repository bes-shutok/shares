from os import PathLike
from pathlib import Path

import openpyxl
from typing import Union

from parsing import parse_data
from supplementary import TradeType, TradeAction, \
    TradeActions, CapitalGainLines, TradeActionList, get_year_month, MonthPartitionedTrades, \
    TradePartsWithinMonth, SortedDateRanges, CapitalGainLineAccumulator, TradeActionsPerCompany, \
    print_month_partitioned_trades, CapitalGainLine, CapitalGainLinesPerCompany


def capital_gains_for_company(trade_actions: TradeActions, symbol: str, currency: str) -> CapitalGainLines:
    capital_gain_line_accumulator = CapitalGainLineAccumulator(symbol, currency)
    sale_trade_parts: TradeActionList = trade_actions[TradeType.SELL]
    buy_trade_parts: TradeActionList = trade_actions[TradeType.BUY]
    if not sale_trade_parts:
        return []
    if not buy_trade_parts:
        raise ValueError("There are sells but no buy_part trades in the provided 'trade_actions' object!")

    sales_monthly_slices = split_by_months(sale_trade_parts, TradeType.SELL)
    buys_monthly_slices = split_by_months(buy_trade_parts, TradeType.BUY)
    capital_gain_lines: CapitalGainLines = []

    while len(sales_monthly_slices) > 0 and len(buys_monthly_slices) > 0:
        sorted_sale_dates_slices: SortedDateRanges = sorted(sales_monthly_slices.keys())
        print("\nsales_monthly_slices:")
        print_month_partitioned_trades(sales_monthly_slices)
        sorted_buy_dates_slices: SortedDateRanges = sorted(buys_monthly_slices.keys())
        print("\nbuys_monthly_slices:")
        print_month_partitioned_trades(buys_monthly_slices)

        # for sale_date in sorted_sale_dates_slices:
        sale_date = sorted_sale_dates_slices[0]
        # for buy_date in sorted_buy_dates_slices:
        buy_date = sorted_buy_dates_slices[0]

        sale_trade_parts: TradePartsWithinMonth = sales_monthly_slices[sale_date]
        print("sale_trade_parts")
        print(sale_trade_parts)

        buy_trade_parts: TradePartsWithinMonth = buys_monthly_slices[buy_date]
        print("buy_trade_parts")
        print(buy_trade_parts)

        target_quantity: int = min(buy_trade_parts.quantity(), sale_trade_parts.quantity())
        sale_quantity_left = target_quantity
        buy_quantity_left = target_quantity
        iteration_count = 0
        while sale_trade_parts.quantity() > 0 and buy_trade_parts.quantity() > 0:
            print("\ncapital_gain_line aggregation cycle (" + str(iteration_count) + ")")
            iteration_count += 1

            extract_trades(sale_quantity_left, sale_trade_parts, capital_gain_line_accumulator)
            extract_trades(buy_quantity_left, buy_trade_parts, capital_gain_line_accumulator)
            print(capital_gain_line_accumulator)

            if sale_trade_parts.count() == 0:
                # remove empty trades
                sales_monthly_slices.pop(sale_date)
            print("sales_monthly_slices")
            print_month_partitioned_trades(sales_monthly_slices)

            if buy_trade_parts.count() == 0:
                # remove empty trades
                buys_monthly_slices.pop(buy_date)
            print("buys_monthly_slices")
            print_month_partitioned_trades(buys_monthly_slices)

        capital_gain_line: CapitalGainLine = capital_gain_line_accumulator.finalize()
        capital_gain_lines.append(capital_gain_line)

    print(capital_gain_lines)

    return capital_gain_lines


def extract_trades(quantity_left, trade_parts, capital_gain_line_accumulator):
    while quantity_left > 0:
        part = trade_parts.pop_trade_part()
        quantity_left -= part[0]
        if quantity_left >= 0:
            capital_gain_line_accumulator.add_trade(part[0], part[1])
        else:
            capital_gain_line_accumulator.add_trade(part[0] + quantity_left, part[1])
            trade_parts.push_trade_part(-quantity_left, part[1])
            quantity_left = 0


def split_by_months(actions: TradeActionList, trade_type: TradeType) -> MonthPartitionedTrades:
    month_partitioned_trades: MonthPartitionedTrades = {}

    if not actions:
        return {}

    for trade_action_part in actions:
        quantity: int = trade_action_part[0]
        trade_action: TradeAction = trade_action_part[1]
        if trade_action.trade_type is not None and trade_action.trade_type != trade_type:
            raise ValueError("Incompatible trade types! Got " + str(trade_type) + "for expected output and " +
                             trade_action.trade_type + " for the trade_action" + str(trade_action))
        year_month = get_year_month(trade_action.date_time)
        trades_within_month: TradePartsWithinMonth = month_partitioned_trades.get(year_month, TradePartsWithinMonth())
        print("pushing trade action " + str(trade_action))
        trades_within_month.push_trade_part(quantity, trade_action)
        month_partitioned_trades[year_month] = trades_within_month

    return month_partitioned_trades


# https://openpyxl.readthedocs.io/en/latest/tutorial.html
def persist_results(path: Union[str, PathLike[str]], capital_gain_lines_per_company: CapitalGainLinesPerCompany):
    first_header = ["Beneficiary", "Country of Source", "SALE", "", "", "PURCHASE", "", "",
                    "WITHOLDING TAX", "", "Expenses incurred with obtaining the capital gains", "",
                    "Symbol", "Currency", "Sale amount", "Buy amount", "Expenses amount"]
    second_header = ["", "", "Month ", "Year", "Amount", "Month ", "Year", "Amount", "Country", "Amount",
                     "", "", "", "", "", "", ""]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporting"

    for i in range(len(first_header)):
        worksheet.cell(1, i + 1, first_header[i])
        worksheet.cell(2, i + 1, second_header[i])

    start_column = 3
    line_number = 3
    for symbol, capital_gain_lines in capital_gain_lines_per_company.items():
        for line in capital_gain_lines:
            worksheet.cell(line_number, start_column, line.get_sell_date().get_month_name())
            worksheet.cell(line_number, start_column + 1, line.get_sell_date().year)
            worksheet.cell(line_number, start_column + 3, line.get_buy_date().get_month_name())
            worksheet.cell(line_number, start_column + 4, line.get_buy_date().year)
            worksheet.cell(line_number, start_column + 10, symbol)
            worksheet.cell(line_number, start_column + 11, line.get_currency())
            worksheet.cell(line_number, start_column + 12, line.get_sell_amount())
            worksheet.cell(line_number, start_column + 13, line.get_buy_amount())
            worksheet.cell(line_number, start_column + 14, line.get_expense_amount())
            line_number += 1

    workbook.save(path)


def main():
    print("Starting conversion.")
    capital_gain_lines_per_company: CapitalGainLinesPerCompany = {}
    trade_actions_per_company: TradeActionsPerCompany = parse_data(Path('resources', 'shares.csv'))
    for symbol, trade_actions in trade_actions_per_company.items():
        if len(trade_actions) != 2:
            continue
        trade_action: TradeAction = trade_actions[TradeType.SELL][0][1]
        assert symbol == trade_action.symbol
        currency = trade_action.currency

        capital_gain_lines: CapitalGainLines = capital_gains_for_company(trade_actions, symbol, currency)
        capital_gain_lines_per_company[symbol] = capital_gain_lines

    xlsx_file = Path('resources', 'tmp.xlsx')
    persist_results(xlsx_file, capital_gain_lines_per_company)


if __name__ == "__main__":
    main()
