import unittest
import pandas as pd
from datetime import datetime

from reporting import split_by_months, create_extract
from supplementary import get_year_month, YearMonth, TradeType, TradePartsWithinMonth, MonthPartitionedTrades, \
    safe_remove_file
from test_common_data import sell_action1

test_dict1 = {("2022", "01"), ("2021", "12"), ("2021", "02")}
test_dict2 = ["202201", "202112", "202102"]
date_time1 = datetime.strptime("2021-05-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
date_time2 = datetime.strptime("2022-05-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
date_time3 = datetime.strptime("2021-01-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
date_time4 = datetime.strptime("2021-12-18, 14:53:23", '%Y-%m-%d, %H:%M:%S')
test_dict4 = [get_year_month(date_time1), get_year_month(date_time2), get_year_month(date_time3),
              get_year_month(date_time4)]
test_dict5 = [YearMonth(date_time1), YearMonth(date_time2), YearMonth(date_time3), YearMonth(date_time4)]


class MyTestCase(unittest.TestCase):

    def test_sorting(self):
        print(str(test_dict1))
        print(str(sorted(test_dict1)))
        print(str(test_dict2))
        print(str(sorted(test_dict2)))
        print(str(test_dict4))
        print(str(sorted(test_dict4)))
        print(test_dict5)
        print(str(sorted(test_dict5)))
        self.assertEqual(1, 1)

    def test_partitioning(self):
        trades_within_month1 = TradePartsWithinMonth()
        trades_within_month1.push_trade_part(1, sell_action1)
        month_partitioned_trades1: MonthPartitionedTrades = {
            get_year_month(sell_action1.date_time): trades_within_month1}

        actual: MonthPartitionedTrades = split_by_months([(1, sell_action1)], TradeType.SELL)
        print(actual)
        self.assertEqual(actual, month_partitioned_trades1)

    # https://kanoki.org/2019/02/26/compare-two-excel-files-for-difference-using-python/
    def test_sorting(self):
        from pathlib import Path

        expected = Path('resources/test', 'extract.xlsx')
        source = Path('resources/test', 'shares.csv')
        destination = Path('resources/test', 'tmp.xlsx')
        destination2 = Path('resources/test', 'tmp2.xlsx')

        safe_remove_file(destination)
        safe_remove_file(destination2)
        create_extract(source, destination)

        from openpyxl import load_workbook
        wb1 = load_workbook(expected, data_only=True)
        wb2 = load_workbook(destination, data_only=True)
        df1 = pd.DataFrame(wb1.active.values)
        df2 = pd.DataFrame(wb2.active.values)

        wb1.save(destination2)

        columns = pd.read_excel(expected,
                                nrows=0,  # Read 0 rows, assuming headers are at row 0
                                ).columns
        str_converter = {col: str for col in columns}  # Convert all fields to strings
        df1 = pd.read_excel(destination, converters=str_converter)
        df2 = pd.read_excel(expected, converters=str_converter)



        wb1.save(destination2)
        from pandas import DataFrame
        df1 = DataFrame(wb1.active.values)
        df2 = DataFrame(wb2.active.values)

        # os.remove(destination)
        comparison_values = df1.values == df2.values
        print(comparison_values)
        self.assertTrue(df1.equals(df2))


if __name__ == '__main__':
    unittest.main()
